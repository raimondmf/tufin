# For opening workbook from xlsx and csv
from openpyxl import load_workbook
# Easy to learn and use template engine, not limited to only html/xml can be any documents
from mako.template import Template
# A module easy to get the subnet mask bit, subnet mask, subnet and ip address
from netaddr import IPNetwork, IPAddress
# For creating xml document
import xml.etree.ElementTree as ET
# For calling REST APIs, very easy to learn and use
import requests
# For using the gethostbyname() method to resolve hostname to IP address
# gethostbyname() does not throw exception when the input is an ip address
# use with caution!
import socket
import re
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import sys
import argparse
parser = argparse.ArgumentParser(description="Crear un ticket de Tufin SC a partir de un fichero excel")
parser.add_argument("ticketid", type=str, help="Numero del ticket ITSM")
parser.add_argument("priority", type=str, help="Valores permitidos: Critical, High, Normal, Low")
parser.add_argument("subject", type=str, help="Asunto del ticket")
args = parser.parse_args()
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Credential for Tufin SC
USERNAME = 'rmiralles'
USERID = '23'
PASSWORD = 'Nextel12345'


# Use to accept the subject string from an external ticketing system
SUBJECT = args.subject
TICKETID = args.ticketid
PRIORITY = args.priority
control = 1
# XLSFILE = TICKETID".xlxs"

for x in ("Critical","High","Normal","Low"):
    if PRIORITY == x:
        control = 0

if control == 1:
    print("\n Send_TufinSC.py TICKETID PRIORITY SUBJECT")
    print(" los valores de PRIORITY solo pueden ser:\n  Critical\n  High\n  Normal\n  Low\n")
    sys.exit(1)

# Reference: Regular Expression Cookbook by Steven Levithan; Jan Goyvaerts
# Published by O'Reilly Media, Inc., 2009
# 7.16. Matching IPv4 Addresses
IPV4_REGEX = "^(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"


# use row=2 if you are using MS Excel, row=3 because I was using Mac Numbers.
row = 2
# Column numbers, you need to adjust according to your excel sheet format
source_id = 1
destination_id = 2
service_id = 3
action_id = 4
comment_id = 5
# Initialize all to zeroes
sources = []
destinations = []
services = []
source_blocks = []
destination_blocks = []
service_blocks = []
access_requests = []



# Collect sources, destinations and services
def get_rule_items(ws, row, column):
    items = []
    while ws.cell(row, column).value is not None:
        items.append(ws.cell(row, column).value)
        row += 1
    return items


# Create source, destination and service xml trees
def construct_rule_items_template(item, item_id):
    if item_id is source_id:
        if "-" in item:
            first_ip = item.split('-')[0]
            second_ip = item.split('-')[1]
            if re.match(IPV4_REGEX, first_ip) and re.match(IPV4_REGEX, second_ip):
                source_block = ET.Element('source')
                source_block.set("type", "RANGE")
                ET.SubElement(source_block, 'range_first_ip').text = first_ip
                ET.SubElement(source_block, 'range_last_ip').text = second_ip
                print(ET.tostring(source_block, encoding='utf-8').decode('utf-8'))
                print(item)
                return source_block
            else:
                try:
                    print(item)
                    answer = socket.gethostbyname(item)
                    source_block = ET.Element('source')
                    source_block.set("type", "IP")
                    ET.SubElement(source_block, 'ip_address').text = str(IPAddress(answer))
                    ET.SubElement(source_block, 'cidr').text = str(IPAddress(answer).netmask_bits())
                    print(ET.tostring(source_block, encoding='utf-8').decode('utf-8'))
                    return source_block
                except Exception as e:
                    print(e)
                    pass
        elif "/" in item:
            try:
                network_id = str(IPNetwork(item).ip)
                network_mask = str(IPNetwork(item).netmask)
                prefix_length = str(IPNetwork(item).prefixlen)
                source_block = ET.Element('source')
                source_block.set("type", "IP")
                ET.SubElement(source_block, 'ip_address').text = network_id
                ET.SubElement(source_block, 'netmask').text = network_mask
                ET.SubElement(source_block, 'cidr').text = prefix_length
                print(ET.tostring(source_block, encoding='utf-8').decode('utf-8'))
                return source_block
            except Exception as e:
                print(e)
                pass
        elif re.match(IPV4_REGEX, item):
            source_block = ET.Element('source')
            source_block.set("type", "IP")
            ET.SubElement(source_block, 'ip_address').text = str(IPAddress(item))
            ET.SubElement(source_block, 'cidr').text = str(IPAddress(item).netmask_bits())
            print(ET.tostring(source_block, encoding='utf-8').decode('utf-8'))
            return source_block
        else:
            answer = socket.gethostbyname(item)
            source_block = ET.Element('source')
            source_block.set("type", "IP")
            ET.SubElement(source_block, 'ip_address').text = str(IPAddress(answer))
            ET.SubElement(source_block, 'cidr').text = str(IPAddress(answer).netmask_bits())
            print(ET.tostring(source_block, encoding='utf-8').decode('utf-8'))
            return source_block
    elif item_id is destination_id:
        if "-" in item:
            first_ip = item.split('-')[0]
            second_ip = item.split('-')[1]
            if re.match(IPV4_REGEX, first_ip) and re.match(IPV4_REGEX, second_ip):
                destination_block = ET.Element('destination')
                destination_block.set("type", "RANGE")
                ET.SubElement(destination_block, 'range_first_ip').text = first_ip
                ET.SubElement(destination_block, 'range_last_ip').text = second_ip
                print(ET.tostring(destination_block, encoding='utf-8').decode('utf-8'))
                print(item)
                return destination_block
            else:
                try:
                    print(item)
                    answer = socket.gethostbyname(item)
                    destination_block = ET.Element('destination')
                    destination_block.set("type", "IP")
                    ET.SubElement(destination_block, 'ip_address').text = str(IPAddress(answer))
                    ET.SubElement(destination_block, 'cidr').text = str(IPAddress(answer).netmask_bits())
                    print(ET.tostring(destination_block, encoding='utf-8').decode('utf-8'))
                    return destination_block
                except Exception as e:
                    print(e)
                    pass
        elif "/" in item:
            try:
                network_id = str(IPNetwork(item).ip)
                network_mask = str(IPNetwork(item).netmask)
                prefix_length = str(IPNetwork(item).prefixlen)
                destination_block = ET.Element('destination')
                destination_block.set("type", "IP")
                ET.SubElement(destination_block, 'ip_address').text = network_id
                ET.SubElement(destination_block, 'netmask').text = network_mask
                ET.SubElement(destination_block, 'cidr').text = prefix_length
                print(ET.tostring(destination_block, encoding='utf-8').decode('utf-8'))
                return destination_block
            except Exception as e:
                print(e)
                pass
        elif re.match(IPV4_REGEX, item):
            destination_block = ET.Element('destination')
            destination_block.set("type", "IP")
            ET.SubElement(destination_block, 'ip_address').text = str(IPAddress(item))
            ET.SubElement(destination_block, 'cidr').text = str(IPAddress(item).netmask_bits())
            print(ET.tostring(destination_block, encoding='utf-8').decode('utf-8'))
            return destination_block
        else:
            answer = socket.gethostbyname(item)
            destination_block = ET.Element('destination')
            destination_block.set("type", "IP")
            ET.SubElement(destination_block, 'ip_address').text = str(IPAddress(answer))
            ET.SubElement(destination_block, 'cidr').text = str(IPAddress(answer).netmask_bits())
            print(ET.tostring(destination_block, encoding='utf-8').decode('utf-8'))
            return destination_block
    elif item_id is service_id:
        service_block = ET.Element('service')
        service_block.set("type", "PROTOCOL")
        if 'tcp' in item.lower():
            ET.SubElement(service_block, 'protocol').text = 'TCP'
            ET.SubElement(service_block, 'port').text = item.lower().split('tcp')[1]
            return service_block
        elif 'udp' in item.lower():
            ET.SubElement(service_block, 'protocol').text = 'UDP'
            ET.SubElement(service_block, 'port').text = item.lower().split('udp')[1]
            return service_block
        else:
            print("Error, the item is not supported.")
    else:
        print("Unknown item, not ip address, subnet or service.")


# Create the access request xml tree
def contruct_access_request_template():
    ar = ET.Element('access_request')
    ET.SubElement(ar, 'use_topology').text = 'true'
    targets = ET.Element('targets')
    target = ET.Element('target')
    target.set("type", "ANY")
    targets.append(target)
    ar.append(targets)
    src = ET.Element('sources')
    dst = ET.Element('destinations')
    svc = ET.Element('services')
    return {'ar': ar, 'sources': src, 'destinations': dst, 'services': svc}


# Prepare for REST API request session
def initialize_requests(username, password):
    tufin_sc = requests.session()
    tufin_sc.headers.update({'Content-type': 'application/xml'})
    tufin_sc.verify = False
    tufin_sc.auth = requests.auth.HTTPBasicAuth(username, password)
    return tufin_sc

# Open the MS Excel workbook
rulebook = load_workbook("rules.xlsx")

# Save the worksheet name for pointing the cells in the correct worksheet
rules = rulebook.sheetnames

if __name__ == '__main__':
    # Initialize the ar fields: source, destination, service
    field = {}
    # Start the template
    template = Template(filename='templates/base.xml')
    # On every worksheet do below
    for rule in rules:
        # Create access request xml for each worksheets
        field = contruct_access_request_template()
        # Assume no more access request if the next worksheet has no data on the data rows
        if rulebook[rule].cell(row, source_id).value is not None or rulebook[rule].cell(row, destination_id).value \
                is not None or rulebook[rule].cell(row, service_id).value is not None:
            # Collect source addresses, destination addresses and services
            sources = get_rule_items(rulebook[rule], row, source_id)
            destinations = get_rule_items(rulebook[rule], row, destination_id)
            services = get_rule_items(rulebook[rule], row, service_id)
            # Collect all source xml blocks
            for source in sources:
                source_blocks.append(construct_rule_items_template(source, source_id))
            # Collect all destination xml blocks
            for destination in destinations:
                destination_blocks.append(construct_rule_items_template(destination, destination_id))
            # Collect all service xml blocks
            for service in services:
                service_blocks.append(construct_rule_items_template(service, service_id))
            # Reconcile the source xml blocks to parent node - sources tag
            for source_block in source_blocks:
                field['sources'].append(source_block)
            # Reconcile the destination xml blocks to parent node - destinations tag
            for destination_block in destination_blocks:
                field['destinations'].append(destination_block)
            # Reconcile the service xml blocks to parent node - services tag
            for service_block in service_blocks:
                field['services'].append(service_block)
        # Reconcile all sources xml blocks, destinations xml blocks and services xml blocks
        # back to grand daddy access request
        field['ar'].append(field['sources'])
        field['ar'].append(field['destinations'])
        field['ar'].append(field['services'])
        # Each access request will have an action field, this one assume Accept.
        # Tufin SC can only do accept. Although Tufin SC has remove and deny option,
        # those are for documentation only, Tufin SC can only remove Cisco ACL (routers)
        # Tufin SC definitely cannot do deny rule.
        ET.SubElement(field['ar'], 'action').text = 'Accept'
        ET.SubElement(field['ar'], 'labels')
        # Initialize all for the next worksheet, to create fresh access request xml block
        sources = []
        destinations = []
        services = []
        source_blocks = []
        destination_blocks = []
        service_blocks = []
        # Collect access request xml blocks
        access_requests.append(ET.tostring(field['ar'], encoding='utf-8').decode('utf-8'))
    # Get the entire xml body from template
    body = template.render(SUBJECT=SUBJECT, access_requests=access_requests, TICKETID=TICKETID, PRIORITY=PRIORITY)
    # Can disable, for me to see the results for troubleshooting only.
    print(body)
    # Prepare the session
    tufin_sc = initialize_requests(USERNAME, PASSWORD)
    # Start calling the REST API POST method, the payload has to be encoded in utf-8
    # payload is byte type which will be sent over to Tufin SC
    try:
        response = tufin_sc.post("https://192.168.13.17/securechangeworkflow/api/securechange/tickets", data=body.encode('utf-8'))
        # Can disable, but for troubleshooting, the response from the server is clearer than postman's response.
        print(response.text)
    except Exception as e:
        print(response.text, e)